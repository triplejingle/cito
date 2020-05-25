import os

import openpyxl
import xlsxwriter


class Excel(object):
    row = 0
    column = 0

    def __init__(self):
        self.document = None
        self.worksheet = None

    def create_document(self, file):
        self.document = xlsxwriter.Workbook(file)

    @staticmethod
    def remove_document(file):
        os.remove(file)

    def save_document(self):
        self.document.close()

    def add_worksheet(self):
        self.worksheet = self.document.add_worksheet()

    def add_data_to_document(self, data):
        for row in data:
            for value in row:
                self.add_cell_data(self.row, self.column, value)
                self.column += 1
            self.column = 0
            self.row += 1

    def add_cell_data(self, row, column, value):
        self.worksheet.write(row, column, value)

    @staticmethod
    def read_cell_data(row, column, file):
        book = openpyxl.load_workbook(file)
        sheet = book.active

        return sheet.cell(row + 1, column + 1).value

    @staticmethod
    def get_document(file):
        return openpyxl.load_workbook(file)
